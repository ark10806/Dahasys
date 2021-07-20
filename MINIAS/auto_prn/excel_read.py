from openpyxl import load_workbook
import xls_path
wb = load_workbook(xls_path.PATH_r, data_only=True)

for sheet in wb:
    print(f'\n\n{sheet}')
    print('-'*5)
    print(sheet.cell(row=1, column=1).value)
    for row in sheet:
        for cell in row:
            print(cell.value)
    
    res = []
    print(sheet['C1'].value)
    exit()

# all_values = []
# for row in ws.rows:
#     tmp = []
#     for cell in row:
#         tmp.append(cell.value)
#     all_values.append(tmp)
# print(all_values)
# load_ws.cell(3, 3, 51470)
# load_ws.cell(4, 3, 21470)
# load_ws.cell(5, 3, 1470)
# load_ws.cell(6, 3, 6470)
# load_wb.save("C:/Users/Administrator/Desktop/기준/프로그래밍/과제대행/주식데이터크롤링/output.xlsx")